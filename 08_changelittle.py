import os
from openpyxl import load_workbook
from datetime import datetime

# 指定目录路径
directory = "./archiv"

def modify_output_file(directory):
    excel_file = "./outputchange.xlsx"

    # 从 "./outputchange.xlsx" 文件的 F2 单元格读取项目名称
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        project_name = sheet["F2"].value
    except FileNotFoundError:
        print(f"{excel_file} 不存在")
        return
    except KeyError:
        print("无法读取项目名称")
        return

    for root, dirs, files in os.walk(directory):
        for folder_name in dirs:
            folder_path = os.path.join(root, folder_name)
            output_file = os.path.join(folder_path, "output.xlsx")
            outputchange_file = os.path.join(folder_path, "outputchange.xlsx")

            # 复制 output.xlsx 文件为 outputchange.xlsx
            try:
                wb = load_workbook(output_file)
                wb.save(outputchange_file)
            except FileNotFoundError:
                print(f"{output_file} 不存在")
                continue

            # 打开 outputchange.xlsx 文件
            try:
                wb = load_workbook(outputchange_file)
                sheet = wb.active
            except FileNotFoundError:
                print(f"{outputchange_file} 不存在")
                continue

            # 获取第3列的截断并删除"-"以及之前的字符，保留后面的内容
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    value = cell.value
                    if value is not None and "-" in value:
                        truncated_value = value.split("-")[-1].strip()
                        cell.value = truncated_value

            # 获取当前年、月、日
            current_year = datetime.now().year
            current_month = datetime.now().month
            current_day = datetime.now().day

            # 备份第一行和最后一行的内容
            first_row_values = [cell.value for cell in sheet[1]]
            last_row_values = [cell.value for cell in sheet[sheet.max_row]]

            # 清除所有行，保留第一行
            sheet.delete_rows(2, sheet.max_row)

            # 在第9、10、11列分别生成当前的年、月、日（带有标题行）
            sheet.cell(row=1, column=9, value="年")
            sheet.cell(row=1, column=10, value="月")
            sheet.cell(row=1, column=11, value="日")

            # 将最后一行的内容移动到第二行位置
            sheet.insert_rows(2)
            for column, value in enumerate(last_row_values, start=1):
                sheet.cell(row=2, column=column).value = value

            # 在第2行写入当前的年、月、日
            sheet.cell(row=2, column=9, value=current_year)
            sheet.cell(row=2, column=10, value=current_month)
            sheet.cell(row=2, column=11, value=current_day)

            # 在第12列写入项目名称，并设置为标题
            sheet.cell(row=1, column=12, value="项目名称")
            sheet.cell(row=2, column=12, value=project_name)

            # 保存修改后的 outputchange.xlsx 文件
            wb.save(outputchange_file)

# 调用函数复制并修改文件
modify_output_file(directory)
