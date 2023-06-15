import os
from openpyxl import Workbook, load_workbook

# 指定目录路径和输出文件路径
directory = "./archiv"
output_file = "./mini/export.xlsx"

def merge_output_files(directory, output_file):
    # 创建一个新的工作簿
    output_workbook = Workbook()

    for root, dirs, files in os.walk(directory):
        for folder_name in dirs:
            folder_path = os.path.join(root, folder_name)
            outputchange_file = os.path.join(folder_path, "outputchange.xlsx")

            # 检查outputchange.xlsx文件是否存在
            if os.path.isfile(outputchange_file):
                try:
                    # 打开outputchange.xlsx文件
                    wb = load_workbook(outputchange_file)
                    sheet = wb.active

                    # 获取除第一行以外的所有行数据
                    data = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        data.append(row)

                    # 复制数据到输出工作簿中
                    output_worksheet = output_workbook.active
                    for row in data:
                        output_worksheet.append(row)

                except FileNotFoundError:
                    print(f"{outputchange_file} 不存在")
                except Exception as e:
                    print(f"处理文件{outputchange_file}时发生错误：{str(e)}")

    # 在第一行插入新的标题行
    output_worksheet = output_workbook.active
    output_worksheet.insert_rows(1)
    titles = [
        "文件编号", "文件题名", "页号", "序号", "责任人", "日期", "页数",
        "档号", "年", "月", "日", "项目名称"
    ]
    for column, title in enumerate(titles, start=1):
        output_worksheet.cell(row=1, column=column, value=title)

    # 保存合并后的输出文件
    output_workbook.save(output_file)
    print(f"合并文件已保存为{output_file}")

# 调用函数合并文件
merge_output_files(directory, output_file)
