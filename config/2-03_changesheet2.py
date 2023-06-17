import os
from openpyxl import Workbook
from openpyxl import load_workbook
from PyPDF2 import PdfFileReader

# 指定目录路径
directory = "../archiv"

def get_pdf_page_count(pdf_path):
    with open(pdf_path, "rb") as file:
        pdf = PdfFileReader(file)
        return pdf.getNumPages()

def write_pdf_names_to_excel(directory):
    # 加载 "../inf.xlsx" 文件
    inf_file = "../inf.xlsx"
    inf_workbook = load_workbook(inf_file)
    inf_sheet = inf_workbook.active
    
    # 从 "../inf.xlsx" 的 B6 单元格中获取责任人
    responsible_person = inf_sheet["B6"].value
    
    for root, dirs, files in os.walk(directory):
        for folder_name in dirs:
            folder_path = os.path.join(root, folder_name)
            excel_file = os.path.join(folder_path, "output.xlsx")
            
            # 从对应位置的 date.txt 文件获取日期
            date_file = os.path.join(directory, folder_name, "date.txt")
            if os.path.exists(date_file):
                with open(date_file, "r", encoding="utf-8") as f:
                    date = f.read().strip()
            else:
                # 输入"日期"
                date = input(f"请输入 {folder_name} 文件夹的日期：")
                # 生成 UTF-8 格式的 date.txt 文件
                os.makedirs(os.path.dirname(date_file), exist_ok=True)
                with open(date_file, "w", encoding="utf-8") as f:
                    f.write(date)
            
            # 创建 Excel 工作簿
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"  # 设置第一个 sheet 的名称为 "Sheet1"
            
            pdf_files = [file for file in os.listdir(folder_path) if file.lower().endswith(".pdf")]
            
            # 将 PDF 文件名写入 Excel 表格的第一列并分列
            for index, pdf_file in enumerate(pdf_files, start=2):  # 从第二行开始写入文件名
                file_name = os.path.splitext(pdf_file)[0]
                separated_names = file_name.split("__")
                for column, name in enumerate(separated_names, start=1):
                    sheet.cell(row=index, column=column).value = name
                
                # 获取 PDF 文件页数并写入第7列
                pdf_path = os.path.join(folder_path, pdf_file)
                page_count = get_pdf_page_count(pdf_path)
                sheet.cell(row=index, column=7).value = page_count
            
            # 在第一行插入新的标题行
            titles = ["文件编号", "文件题名", "页号", "序号", "责任人", "日期", "页数","档号"]
            for column, title in enumerate(titles, start=1):
                sheet.cell(row=1, column=column).value = title
            
            # 自动排序序号列
            for index, row in enumerate(sheet.iter_rows(min_row=2, min_col=4, max_col=4), start=1):
                for cell in row:
                    cell.value = index
            
            # 根据页数在第3列添加数据
            start_page = 1
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    page_count = cell.offset(column=4).value
                    if page_count is None:
                        break
                    end_page = start_page + page_count - 1
                    cell.value = f"{start_page}-{end_page}"
                    start_page = end_page + 1
            
            # 填写"责任人"和"日期"到第5列和第6列
            for index in range(2, sheet.max_row + 1):
                sheet.cell(row=index, column=5).value = responsible_person
                sheet.cell(row=index, column=6).value = date
            
            # 在第8列，写入截断后的文件名
            for index in range(2, sheet.max_row + 1):
                file_number = sheet.cell(row=index, column=1).value
                truncated_number = file_number[:-4]
                sheet.cell(row=index, column=8).value = truncated_number
            
            # 保存 Excel 文件
            workbook.save(excel_file)

# 调用函数将 PDF 文件名写入 Excel 表格
write_pdf_names_to_excel(directory)
