import os
from openpyxl import Workbook, load_workbook

# 指定目录路径和 Excel 文件路径
directory = "../archiv"
excel_file = "./outputchange.xlsx"

def write_folder_names_to_excel(directory, excel_file):
    # 创建 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"  # 设置第一个 sheet 的名称为 "Sheet1"

    # 在 Sheet1 的第一行插入数据
    sheet.insert_rows(1)
    sheet["A1"] = "档号"
    sheet["B1"] = "案卷题名"
    sheet["C1"] = "序号"
    sheet["D1"] = "份数"
    sheet["E1"] = "保管期限"
    sheet["F1"] = "项目名称"
    sheet["G1"] = "公司名称"
    sheet["H1"] = "单位性质"

    # 获取指定目录下的所有文件夹名称
    folder_names = [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]

    # 从 "../inf.xlsx" 的 B1、B4、B5、B2 和 B3 单元格中获取项目名称、份数、保管期限、公司名称和单位性质
    inf_file = "../inf.xlsx"
    inf_workbook = load_workbook(inf_file)
    inf_sheet = inf_workbook.active
    project_name = inf_sheet["B1"].value
    file_count = inf_sheet["B4"].value
    retention_period = inf_sheet["B5"].value
    company_name = inf_sheet["B2"].value
    unit_property = inf_sheet["B3"].value

    # 将文件夹名称写入 Excel 表格的第一列并分列
    for index, folder_name in enumerate(folder_names, start=2):
        separated_names = folder_name.split("__")
        for column, name in enumerate(separated_names, start=1):
            sheet.cell(row=index, column=column).value = name

        # 在第3列插入序号
        sheet.cell(row=index, column=3).value = index - 1

        # 在第4列插入标题份数
        sheet.cell(row=index, column=4).value = file_count

        # 在第5列插入保管期限
        sheet.cell(row=index, column=5).value = retention_period

        # 在第6列插入项目名称
        sheet.cell(row=index, column=6).value = project_name
        
        # 在第7列插入公司名称
        sheet.cell(row=index, column=7).value = company_name
        
        # 在第8列插入单位性质
        sheet.cell(row=index, column=8).value = unit_property

    # 保存 Excel 文件
    workbook.save(excel_file)

# 调用函数将文件夹名称写入 Excel 表格
write_folder_names_to_excel(directory, excel_file)
