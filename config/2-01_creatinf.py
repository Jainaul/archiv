import os
import openpyxl

file_path = "../inf.xlsx"
print_message = True

if os.path.exists(file_path):
    overwrite = input(f"文件 {file_path} 已存在，是否覆盖？ (y/n): ").lower()
    if overwrite == "n":
        print("将使用已有信息生成文件！")
        print_message = False
    elif overwrite == "y":
        # 打开现有的Excel文件
        workbook = openpyxl.load_workbook(file_path)

        # 获取默认的工作表
        sheet = workbook.active

        # 写入内容到第一列和第二列
        content = [
            ["项目名称：", "XXXXX项目"],
            ["公司名称：", "XXXXX设计公司"],
            ["单位性质：", "设计"],
            ["保存份数：", "2"],
            ["保管期限：", "30年"],
            ["责任人：", "张三"]
        ]

        for index, values in enumerate(content, start=1):
            for col, value in enumerate(values, start=1):
                sheet.cell(row=index, column=col, value=value)

        # 保存Excel文件
        workbook.save(file_path)

        print(f"已成功创建并覆盖现有文件 {file_path}")
else:
    # 创建新的Excel工作簿
    workbook = openpyxl.Workbook()

    # 获取默认的工作表
    sheet = workbook.active

    # 写入内容到第一列和第二列
    content = [
        ["项目名称：", "XXXXX项目"],
        ["公司名称：", "XXXXX设计公司"],
        ["单位性质：", "设计"],
        ["保存份数：", "2"],
        ["保管期限：", "30年"],
        ["责任人：", "张三"]
    ]

    for index, values in enumerate(content, start=1):
        for col, value in enumerate(values, start=1):
            sheet.cell(row=index, column=col, value=value)

    # 保存Excel文件
    workbook.save(file_path)

    print(f"已成功创建文件 {file_path}")

if print_message:
    print("请在inf.xlsx的第二列中填入相应信息！")
    input("请在填写完成后按回车键继续！")
