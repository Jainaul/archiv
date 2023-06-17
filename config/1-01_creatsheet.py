import os
from openpyxl import Workbook, load_workbook

# 指定目录路径和 Excel 文件路径
directory = "../archiv"
excel_file = "./output.xlsx"

def write_folder_names_to_excel(directory, excel_file):
    # 创建 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"  # 设置第一个 sheet 的名称为 "Sheet1"
    
    # 获取指定目录下的所有文件夹名称
    folder_names = [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]
    
    # 将文件夹名称写入 Excel 表格的第一列并分列
    for index, folder_name in enumerate(folder_names, start=1):
        separated_names = folder_name.split("__")
        for column, name in enumerate(separated_names, start=1):
            sheet.cell(row=index, column=column).value = name
            print(f"正在处理文件夹: {name}")  # 显示当前正在处理的文件夹
    
    # 保存 Excel 文件
    workbook.save(excel_file)

# 调用函数将文件夹名称写入 Excel 表格
write_folder_names_to_excel(directory, excel_file)
    
def write_file_names_to_excel(directory, excel_file):
    # 加载现有的 Excel 文件
    workbook = load_workbook(excel_file)

    # 创建一个新的 Sheet，命名为 "Sheet2"
    sheet2 = workbook.create_sheet("Sheet2")

    # 递归获取指定目录下的所有文件的文件名
    file_names = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            file_names.append(file_name)
            print(f"正在处理文件: {file_name}")  # 显示当前正在处理的文件
    
    # 将文件名写入 Sheet2 的第一列并分列
    for index, file_name in enumerate(file_names, start=1):
        separated_names = file_name.split("__")
        for column, name in enumerate(separated_names, start=1):
            sheet2.cell(row=index, column=column).value = name

    # 保存 Excel 文件
    workbook.save(excel_file)

# 调用函数将文件夹名称写入 Excel 表格
write_file_names_to_excel(directory, excel_file)
