import os
from openpyxl import Workbook, load_workbook

# 指定目录路径和输出文件路径
directory = "./archiv"
output_file = "./minor/export.xlsx"

def merge_output_files(directory, output_file):
    # 创建一个新的工作簿
    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "Sheet1"

    # 储存所有数据的列表
    data = []

    for root, dirs, files in os.walk(directory):
        for folder_name in dirs:
            folder_path = os.path.join(root, folder_name)
            output_file_path = os.path.join(folder_path, "output.xlsx")

            # 检查output.xlsx文件是否存在
            if os.path.isfile(output_file_path):
                try:
                    # 打开output.xlsx文件
                    wb = load_workbook(output_file_path)
                    sheet = wb.active

                    # 获取2-16行的数据
                    for row in sheet.iter_rows(min_row=2, max_row=16, values_only=True):
                        data.append(row)

                except FileNotFoundError:
                    print(f"{output_file_path} 不存在")
                except Exception as e:
                    print(f"处理文件{output_file_path}时发生错误：{str(e)}")

    # 在第一行插入新的标题行
    titles = [
        "文件编号", "文件题名", "页号", "序号", "责任人", "日期", "页数", "档号"
    ]
    output_worksheet.append(titles)

    # 复制数据到输出工作表中
    for row in data:
        output_worksheet.append(row)

    # 保存合并后的输出文件
    output_workbook.save(output_file)
    print(f"合并文件已保存为{output_file}")

# 调用函数合并文件
merge_output_files(directory, output_file)
