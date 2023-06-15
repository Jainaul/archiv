import os
from openpyxl import Workbook, load_workbook

# 指定目录路径和 Excel 文件路径
directory = "./archiv"
excel_file = "./outputchange.xlsx"

def write_folder_names_to_excel(directory, excel_file):
    # 创建 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"  # 设置第一个 sheet 的名称为 "Sheet1"

    # 在 Sheet1 的第一行插入数据
    sheet.insert_rows(1)
    sheet["A1"] = "档号"
    sheet["B1"] = "案卷题名"
    sheet["C1"] = "序号"
    sheet["D1"] = "份数"
    sheet["E1"] = "保管期限"
    sheet["F1"] = "项目名称"

    # 获取指定目录下的所有文件夹名称
    folder_names = [name for name in os.listdir(directory) if os.path.isdir(os.path.join(directory, name))]

    # 在第5列插入项目名称
    project_name = input("请输入项目名称：")  # 要求用户输入项目名称

    # 将文件夹名称写入 Excel 表格的第一列并分列
    for index, folder_name in enumerate(folder_names, start=2):
        separated_names = folder_name.split("__")
        for column, name in enumerate(separated_names, start=1):
            sheet.cell(row=index, column=column).value = name

        # 在第3列插入序号
        sheet.cell(row=index, column=3).value = index - 1

        # 在第4列插入标题份数（全为"2"）
        sheet.cell(row=index, column=4).value = 2

        # 在第5列插入保管期限（全为"永久"）
        sheet.cell(row=index, column=5).value = "永久"

        # 在第5列插入项目名称
        sheet.cell(row=index, column=6).value = project_name

    # 保存 Excel 文件
    workbook.save(excel_file)

# 调用函数将文件夹名称写入 Excel 表格
write_folder_names_to_excel(directory, excel_file)
